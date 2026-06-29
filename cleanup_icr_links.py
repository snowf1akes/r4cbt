#!/usr/bin/env python3
"""
cleanup_icr_links.py

Two jobs for finalized ICR annotation sheets:

  FIX  — sheets that HAVE a links column but whose OP/Title-row links point to
         the wrong Reddit post (the generator's stale-`post` bug). The correct
         post link is recovered from the post's own comment permalinks.

  ADD  — sheets that have NO links column at all. A new "Link" column is
         appended at the end and populated by matching each post/comment, BY
         CONTENT, to a sibling sheet in the same workbook that does have links
         (e.g. an "ICR" tab). Post links go on the Title row; comment permalinks
         go on each commenter's row where available.

Matching is by content (post title text; author + comment text), never by row
position, because the link-bearing and link-less sheets can drift out of
alignment by a few rows.

Nothing else is touched: annotations, dropdowns, formatting and merged cells are
preserved, and originals are never overwritten (output is <name>_fixed.xlsx).
Every post that can't be resolved is reported as a WARNING rather than guessed.

Usage:
  python cleanup_icr_links.py sheet.xlsx [more.xlsx ...]
  python cleanup_icr_links.py --dir folder/
"""
import argparse
import os
import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

REDDIT_URL = re.compile(
    r"^(https?://(?:www\.)?reddit\.com/r/[^/]+/comments/([a-z0-9]+)(?:/[^/]+)?/?)"
    r"(?:([a-z0-9]+)/?)?$",
    re.IGNORECASE,
)
TITLE_RE = re.compile(r"^\s*Title\s*:", re.IGNORECASE)
KEY_LEN = 25  # chars of comment text used for content matching


def parse_link(url):
    if not url:
        return None, None
    m = REDDIT_URL.match(url.strip())
    return (m.group(2), m.group(3)) if m else (None, None)


def post_url_from_comment(url):
    s = url.strip().rstrip("/")
    return s[: s.rfind("/")] + "/"


def norm(v):
    return " ".join(str(v).split()) if v is not None else ""


def detect_link_col(ws):
    counts = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink and "reddit.com/r/" in (cell.hyperlink.target or ""):
                counts[cell.column] = counts.get(cell.column, 0) + 1
    return max(counts, key=counts.get) if counts else None


def detect_comment_col(ws):
    """Column holding the 'Title:' / comment text (scan the top of the sheet)."""
    for col in range(1, ws.max_column + 1):
        for r in range(1, min(ws.max_row, 80) + 1):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and TITLE_RE.match(v):
                return col
    return 3


def title_rows(ws, ccol):
    return [r for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, ccol).value, str)
            and TITLE_RE.match(ws.cell(r, ccol).value)]


# ---------------------------------------------------------------- FIX mode
def fix_sheet(ws):
    lcol = detect_link_col(ws)
    if lcol is None:
        return None
    ccol = detect_comment_col(ws)
    titles = title_rows(ws, ccol)
    if not titles:
        return None
    bounds = titles + [ws.max_row + 1]
    fixes = ok = 0
    warnings = []
    for i, tr in enumerate(titles):
        op = ws.cell(tr, lcol)
        op_t = op.hyperlink.target if op.hyperlink else None
        op_pid, op_cid = parse_link(op_t)
        pids, sample = set(), None
        for r in range(bounds[i], bounds[i + 1]):
            if r == tr:
                continue
            cell = ws.cell(r, lcol)
            t = cell.hyperlink.target if cell.hyperlink else None
            pid, cid = parse_link(t)
            if pid and cid:
                pids.add(pid)
                sample = t
        if not pids:
            warnings.append(f"row {tr}: no comment links to verify OP link — left as is.")
            continue
        if len(pids) > 1:
            warnings.append(f"row {tr}: comment links disagree {sorted(pids)} — left as is.")
            continue
        correct = post_url_from_comment(sample)
        cpid = next(iter(pids))
        if op_t is None:
            op.value, op.hyperlink = correct, correct
            op.font = Font(color="0000EE", underline="single")
            fixes += 1
        elif op_cid is not None:
            warnings.append(f"row {tr}: OP link is a comment permalink — left as is.")
        elif op_pid == cpid:
            ok += 1
        else:
            op.hyperlink.target = correct
            if isinstance(op.value, str) and "reddit.com" in op.value:
                op.value = correct
            op.font = Font(color="0000EE", underline="single")
            fixes += 1
    return ("fix", ws.title, fixes, ok, warnings)


# ---------------------------------------------------------------- ADD mode
def build_source_index(wb):
    """Map post-title -> {op_post_url, comments:{(author,line25):url}} from
    every sheet that already carries links."""
    posts = {}
    for ws in wb.worksheets:
        lcol = detect_link_col(ws)
        if lcol is None:
            continue
        ccol = detect_comment_col(ws)
        titles = title_rows(ws, ccol)
        bounds = titles + [ws.max_row + 1]
        for i, tr in enumerate(titles):
            title = norm(ws.cell(tr, ccol).value)
            comments, pids, sample = {}, set(), None
            for r in range(bounds[i], bounds[i + 1]):
                if r == tr:
                    continue
                cell = ws.cell(r, lcol)
                if not cell.hyperlink:
                    continue
                pid, cid = parse_link(cell.hyperlink.target)
                if pid and cid:
                    key = (norm(ws.cell(r, 2).value),
                           norm(ws.cell(r, ccol).value)[:KEY_LEN])
                    comments[key] = cell.hyperlink.target
                    pids.add(pid)
                    sample = cell.hyperlink.target
            op_url = post_url_from_comment(sample) if (len(pids) == 1 and sample) else None
            slot = posts.setdefault(title, {"op_post_url": None, "comments": {}})
            slot["comments"].update(comments)
            if op_url and not slot["op_post_url"]:
                slot["op_post_url"] = op_url
    return posts


def add_links_sheet(ws, src):
    if detect_link_col(ws) is not None:
        return None  # already has links — FIX handles it
    ccol = detect_comment_col(ws)
    titles = title_rows(ws, ccol)
    if not titles:
        return None
    new_col = ws.max_column + 1
    hdr = ws.cell(1, new_col, "Link")
    hdr.font = Font(bold=True)
    bounds = titles + [ws.max_row + 1]
    op_done = op_miss = cmt_done = 0
    warnings = []
    for i, tr in enumerate(titles):
        title = norm(ws.cell(tr, ccol).value)
        sp = src.get(title)
        if sp and sp["op_post_url"]:
            c = ws.cell(tr, new_col)
            c.value, c.hyperlink = sp["op_post_url"], sp["op_post_url"]
            c.font = Font(color="0000EE", underline="single")
            op_done += 1
        else:
            op_miss += 1
            warnings.append(f"row {tr}: no source post link for "
                            f"{title[:45]!r} — Title row left blank.")
        if not sp:
            continue
        for r in range(bounds[i], bounds[i + 1]):
            if r == tr or not norm(ws.cell(r, 2).value):
                continue
            key = (norm(ws.cell(r, 2).value), norm(ws.cell(r, ccol).value)[:KEY_LEN])
            url = sp["comments"].get(key)
            if url:
                c = ws.cell(r, new_col)
                c.value, c.hyperlink = url, url
                c.font = Font(color="0000EE", underline="single")
                cmt_done += 1
    return ("add", ws.title, op_done, op_miss, cmt_done, get_column_letter(new_col), warnings)


# ---------------------------------------------------------------- driver
def process_file(path):
    wb = load_workbook(path)
    src = build_source_index(wb)
    fixed_total = added_total = 0
    for ws in wb.worksheets:
        if detect_link_col(ws) is not None:
            res = fix_sheet(ws)
            if res:
                _, name, fixes, ok, warns = res
                fixed_total += fixes
                print(f"  [FIX {name}] {fixes} OP links corrected, {ok} already correct")
                for w in warns:
                    print(f"      WARNING {w}")
        else:
            res = add_links_sheet(ws, src)
            if res:
                _, name, opd, opm, cmt, col, warns = res
                added_total += opd + cmt
                print(f"  [ADD {name}] new col {col}: {opd} post links + {cmt} "
                      f"comment links ({opm} posts had no source link)")
                for w in warns:
                    print(f"      WARNING {w}")
    out = os.path.splitext(path)[0] + "_fixed.xlsx"
    wb.save(out)
    print(f"{os.path.basename(path)}: {fixed_total} fixed, {added_total} added "
          f"-> {os.path.basename(out)}")


def main():
    ap = argparse.ArgumentParser(description="Fix/add Reddit links in ICR sheets")
    ap.add_argument("files", nargs="*")
    ap.add_argument("--dir", help="process every .xlsx in this folder")
    args = ap.parse_args()
    targets = list(args.files)
    if args.dir:
        targets += [os.path.join(args.dir, f) for f in sorted(os.listdir(args.dir))
                    if f.endswith(".xlsx") and not f.endswith("_fixed.xlsx")]
    if not targets:
        ap.print_help()
        sys.exit(1)
    for t in targets:
        print(f"\nProcessing {t}")
        process_file(t)


if __name__ == "__main__":
    main()
